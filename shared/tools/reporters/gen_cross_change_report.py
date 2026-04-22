import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path
_ROOT = Path(__file__).resolve().parent.parent.parent.parent
sys.path.insert(0, str(_ROOT / 'shared' / 'tools'))
from excel_styles import to_openpyxl_styles
import os
from datetime import datetime

# === Load data ===
conn = sqlite3.connect(str(_ROOT / 'projects' / 'ecr-ecn' / 'workspace' / 'db' / 'ecr_ecn.db'))
cursor = conn.cursor()

# Change 1: BSOB (O-0001) — 34 parts, all V BSOB
cursor.execute('''
    SELECT p.part_number, p.product_name, p.product_package, p.product_function,
           p.involvement, p.au_type,
           b.bsob, b.die_size_mil2, b.wire_diameter_mil, b.full_pkg_code
    FROM std_ecr_parts p
    LEFT JOIN std_bsob_detail b ON p.part_number = b.part_number
    WHERE p._source_file = 'O-0001'
''')
change1_rows = cursor.fetchall()
change1_parts = set(r[0] for r in change1_rows if r[0])
change1_dict = {r[0]: r for r in change1_rows if r[0]}

# Change 2: Au->Cu (O-0002) — 只取 involvement='V' (304 parts)
# ECR-N1: 只有「涉及與否=V」才是實際變更範圍
cursor.execute('''
    SELECT p.part_number, p.product_name, p.product_package, p.product_function,
           p.involvement, p.au_type,
           d.eval_batch, d.has_bom, d.bop_gold, d.bop_silver, d.bop_copper, d.wire_type
    FROM std_ecr_parts p
    LEFT JOIN std_au_cu_detail d ON p.part_number = d.part_number
    WHERE p._source_file = 'O-0002' AND p.involvement = 'V'
''')
change2_rows = cursor.fetchall()
change2_parts = set(r[0] for r in change2_rows if r[0])
change2_dict = {r[0]: r for r in change2_rows if r[0]}

# 同時查詢 O-0002 全量（用於變更1關係分析）
cursor.execute('''
    SELECT p.part_number, p.involvement, p.au_type
    FROM std_ecr_parts p
    WHERE p._source_file = 'O-0002'
''')
change2_all_rows = cursor.fetchall()
change2_all_dict = {r[0]: {'involvement': r[1], 'au_type': r[2]} for r in change2_all_rows if r[0]}

conn.close()

# Change 3: Back Au->Sn/Ag from Excel — 791 unique parts, all in scope (ECR-R22)
wb_src = openpyxl.load_workbook(
    str(_ROOT / 'projects' / 'ecr-ecn' / 'vault' / 'originals' / '20260203 Phenitec背金轉背錫 晶片+產品 list.xlsx'),
    read_only=True, data_only=True
)
ws_src = wb_src[wb_src.sheetnames[0]]

change3_parts = set()
change3_dict = {}
for i, row in enumerate(ws_src.iter_rows(values_only=True)):
    if i <= 1:
        continue
    part = str(row[10]).strip() if row[10] else None
    if part and part != 'None':
        change3_parts.add(part)
        if part not in change3_dict:
            change3_dict[part] = {
                'supplier': str(row[0]) if row[0] else '',
                'wafer_type': str(row[1]) if row[1] else '',
                'die_size': str(row[2]) if row[2] else '',
                'die_size_val': row[3],
                'die_size_class': str(row[4]) if row[4] else '',
                'die_thick': row[5],
                'wafer_code': str(row[7]) if row[7] else '',
                'back_side': str(row[8]) if row[8] else '',
                'function': str(row[9]) if row[9] else '',
                'part_number': part,
                'package': str(row[11]) if row[11] else '',
                'ship_2024': row[12] if row[12] else 0,
                'ship_2025': row[13] if row[13] else 0,
                'ship_total': row[14] if row[14] else 0,
                'au_con': str(row[15]) if row[15] else '',
                'lead_frame': str(row[16]) if row[16] else '',
                'epoxy': str(row[17]) if row[17] else '',
                'eutectic': str(row[18]) if row[18] else '',
                'die_attach': str(row[19]) if row[19] else '',
                'sn_ag': str(row[20]) if row[20] else '',
                'npi_command': str(row[21]) if row[21] else '',
                'pkg_code': str(row[23]) if len(row) > 23 and row[23] else '',
            }
wb_src.close()

# === Compute sets ===
# 注意：change2_parts 現在只含 V 的 304 筆
c1_c2 = change1_parts & change2_parts       # 應為 0（BSOB 34筆在 O-0002 中都是 X）
c2_c3 = change2_parts & change3_parts       # 23 筆
c1_c3 = change1_parts & change3_parts       # 0 筆
c1_c2_c3 = change1_parts & change2_parts & change3_parts  # 0 筆
c1_only = change1_parts - change2_parts - change3_parts    # 34 筆
c2_only = change2_parts - change1_parts - change3_parts    # 281 筆
c3_only = change3_parts - change1_parts - change2_parts    # 768 筆
all_union = change1_parts | change2_parts | change3_parts  # 1,106 筆

print(f'=== 跨變更交叉分析（修正版）===')
print(f'變更1 (BSOB):      {len(change1_parts)} parts')
print(f'變更2 (Cu線, V only): {len(change2_parts)} parts')
print(f'變更3 (背錫銀):     {len(change3_parts)} parts')
print(f'---')
print(f'變更1 ∩ 變更2(V):   {len(c1_c2)} parts')
print(f'變更2(V) ∩ 變更3:   {len(c2_c3)} parts')
print(f'變更1 ∩ 變更3:      {len(c1_c3)} parts')
print(f'三者全交集:          {len(c1_c2_c3)} parts')
print(f'---')
print(f'僅變更1:             {len(c1_only)} parts')
print(f'僅變更2(V):          {len(c2_only)} parts')
print(f'僅變更3:             {len(c3_only)} parts')
print(f'總計不重複:          {len(all_union)} parts')

# === Styles ===
_S = to_openpyxl_styles()
thin_border = _S['border']
# Report-specific fills (per-change color coding)
hdr_font = Font(bold=True, color='FFFFFF', size=11)
hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
hdr_fill2 = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
hdr_fill3 = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
hdr_fill4 = PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid')
hdr_fill5 = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
overlap_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

def style_header(ws, row_num, fill, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def auto_width(ws, max_col, max_row):
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 100)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

# === Create Excel ===
wb = openpyxl.Workbook()

# ============ Sheet 1: 總覽 ============
ws1 = wb.active
ws1.title = '總覽'

ws1.merge_cells('A1:F1')
title_cell = ws1.cell(row=1, column=1, value='ECR/ECN 跨變更交叉分析報告（修正版）')
title_cell.font = Font(bold=True, size=14)

ws1.cell(row=2, column=1, value=f'產出日期：{datetime.now().strftime("%Y-%m-%d %H:%M")}')
ws1.cell(row=2, column=1).font = Font(italic=True, color='666666')

# 重要說明
ws1.merge_cells('A3:F3')
note_cell = ws1.cell(row=3, column=1,
    value='重要：變更2(金線轉銅線)僅計算「涉及與否=V」的 304 筆（非全量 7,765 筆）。變更3(Phenitec背金轉背錫)全部 791 料號皆涉及。')
note_cell.font = Font(bold=True, color='CC0000', size=10)
note_cell.fill = note_fill

# 各變更案規模
ws1.cell(row=5, column=1, value='一、各變更案規模').font = Font(bold=True, size=12)
h1 = ['變更案', '方向', '對象', '供應商', '涉及料號數', '備註']
for j, h in enumerate(h1, 1):
    ws1.cell(row=6, column=j, value=h)
style_header(ws1, 6, hdr_fill, len(h1))

data1 = [
    ['變更1 (BSOB)', '金線 → 鈀金銅線', 'Wire bonding', '多家', len(change1_parts), '全部 V BSOB'],
    ['變更2 (Cu線)', '金線 → 銅線', 'Wire bonding', '多家(岡山廠)', len(change2_parts), f'V={len(change2_parts)}（全量 7,765 中）'],
    ['變更3 (背錫銀)', '背金 → 背錫/背銀', 'Die Attach', 'Phenitec', len(change3_parts), '全部涉及（無涉及與否欄位）'],
    ['聯集（不重複）', '', '', '', len(all_union), ''],
]
for i, row in enumerate(data1, 7):
    for j, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.border = thin_border
        if row[0].startswith('聯集'):
            c.font = Font(bold=True)

# 交集統計
ws1.cell(row=12, column=1, value='二、交集統計').font = Font(bold=True, size=12)
h2 = ['交集', '料號數', '佔比', '說明']
for j, h in enumerate(h2, 1):
    ws1.cell(row=13, column=j, value=h)
style_header(ws1, 13, hdr_fill2, len(h2))

pct_c2_c3 = f'{len(c2_c3)/max(len(change2_parts),1)*100:.1f}% of 變更2(V)'
pct_c2_c3_of3 = f'{len(c2_c3)/max(len(change3_parts),1)*100:.1f}% of 變更3'
data2 = [
    ['變更1 ∩ 變更2(V)', len(c1_c2), '0%', 'BSOB 34筆在O-0002中全部標X/BSOB → 不重疊（獨立變更案）'],
    ['變更1 ∩ 變更3', len(c1_c3), '0%', '無重疊'],
    ['變更2(V) ∩ 變更3', len(c2_c3), f'{pct_c2_c3} / {pct_c2_c3_of3}', '同時需改線材 + 改背面金屬'],
    ['三者全交集', len(c1_c2_c3), '-', '無'],
    ['任何跨變更', len(c1_c2 | c1_c3 | c2_c3), '', f'全部為變更2(V)∩3 的 {len(c2_c3)} 筆'],
]
for i, row in enumerate(data2, 14):
    for j, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.border = thin_border

# Venn 分佈
ws1.cell(row=20, column=1, value='三、Venn 圖分佈').font = Font(bold=True, size=12)
h3 = ['區塊', '料號數', '佔比', '說明']
for j, h in enumerate(h3, 1):
    ws1.cell(row=21, column=j, value=h)
style_header(ws1, 21, hdr_fill3, len(h3))

data3 = [
    ['僅變更1 (BSOB獨有)', len(c1_only), f'{len(c1_only)/max(len(all_union),1)*100:.1f}%', 'BSOB 與其他兩案完全獨立'],
    ['僅變更2(V) (Cu線獨有)', len(c2_only), f'{len(c2_only)/max(len(all_union),1)*100:.1f}%', '僅需改線材（金→銅）'],
    ['僅變更3 (背錫銀獨有)', len(c3_only), f'{len(c3_only)/max(len(all_union),1)*100:.1f}%', '僅需改背面金屬（金→錫/銀）'],
    ['變更1∩2(V) (不含3)', len(c1_c2 - c1_c2_c3), '-', 'BSOB在O-0002中為X → 0筆'],
    ['變更2(V)∩3 (不含1)', len(c2_c3 - c1_c2_c3), f'{len(c2_c3 - c1_c2_c3)/max(len(all_union),1)*100:.1f}%', '同時改線材+改背面金屬'],
    ['變更1∩3 (不含2)', len(c1_c3 - c1_c2_c3), '-', ''],
    ['變更1∩2(V)∩3', len(c1_c2_c3), '-', ''],
    ['總計', len(all_union), '100%', ''],
]
for i, row in enumerate(data3, 22):
    for j, val in enumerate(row, 1):
        c = ws1.cell(row=i, column=j, value=val)
        c.border = thin_border
        if row[0] == '總計':
            c.font = Font(bold=True)

# Key findings
ws1.cell(row=31, column=1, value='四、關鍵發現').font = Font(bold=True, size=12)
findings = [
    '1. 三個變更案幾乎完全獨立：僅 23 筆料號跨變更2(V)與變更3，佔總量 2.1%。',
    '2. 變更1(BSOB)完全獨立：34筆料號在O-0002中全部標X/BSOB，不在Cu線變更範圍內。',
    f'3. 變更2(V)與變更3的交集(23筆)需同時進行：線材變更(Au→Cu) + 背面金屬變更(Au→Sn/Ag)。',
    f'4. 變更3佔比最大(768/1,106=69.4%)，其中 791 料號全部涉及（ECR-R22）。',
    '5. 製程搭配限制(ECR-R23)：Epoxy→只能Ag背銀；Eutectic→只能Sn背錫；',
    '   若一顆AU晶片同時有Epoxy和Eutectic製程需求→必須拆成兩種晶片。',
    f'6. 注意：變更2全量為 7,765 筆，但僅 V={len(change2_parts)} 筆實際涉及（ECR-N1）。',
]
for i, f in enumerate(findings, 32):
    ws1.cell(row=i, column=1, value=f)

auto_width(ws1, 6, 40)

# ============ Sheet 2: 變更2(V)∩3 ============
ws2 = wb.create_sheet(f'變更2(V)∩3 ({len(c2_c3)}筆)')
headers = [
    '料號', '品名', '封裝(Cu線)', 'Function',
    'Cu線-涉及與否', 'Cu線-AU類型', 'Cu線-評估批次', 'Cu線-有BOM',
    'Cu線-BOP金', 'Cu線-BOP銀', 'Cu線-BOP銅', 'Cu線-線材',
    '背錫銀-Die Attach', '背錫銀-處置(Sn/Ag)', '背錫銀-NPI Command',
    '背錫銀-封裝', '背錫銀-Au/Con', '背錫銀-Lead Frame',
    '背錫銀-Epoxy', '背錫銀-Eutectic',
    '背錫銀-2024出貨(K)', '背錫銀-2025出貨(K)', '背錫銀-出貨加總'
]
for j, h in enumerate(headers, 1):
    ws2.cell(row=1, column=j, value=h)
style_header(ws2, 1, hdr_fill, len(headers))

row_num = 2
for p in sorted(c2_c3):
    c2 = change2_dict.get(p, (None,)*12)
    c3 = change3_dict.get(p, {})
    vals = [
        p,
        c2[1] if len(c2) > 1 and c2[1] else '',
        c2[2] if len(c2) > 2 and c2[2] else '',
        c2[3] if len(c2) > 3 and c2[3] else '',
        c2[4] if len(c2) > 4 and c2[4] else '',
        c2[5] if len(c2) > 5 and c2[5] else '',
        c2[6] if len(c2) > 6 and c2[6] else '',
        c2[7] if len(c2) > 7 and c2[7] else '',
        c2[8] if len(c2) > 8 and c2[8] else '',
        c2[9] if len(c2) > 9 and c2[9] else '',
        c2[10] if len(c2) > 10 and c2[10] else '',
        c2[11] if len(c2) > 11 and c2[11] else '',
        c3.get('die_attach', ''),
        c3.get('sn_ag', ''),
        c3.get('npi_command', ''),
        c3.get('package', ''),
        c3.get('au_con', ''),
        c3.get('lead_frame', ''),
        c3.get('epoxy', ''),
        c3.get('eutectic', ''),
        c3.get('ship_2024', 0),
        c3.get('ship_2025', 0),
        c3.get('ship_total', 0),
    ]
    for j, val in enumerate(vals, 1):
        c = ws2.cell(row=row_num, column=j, value=val)
        c.border = thin_border
    row_num += 1

auto_width(ws2, len(headers), row_num)
ws2.freeze_panes = 'A2'

# ============ Sheet 3: 變更1與變更2關係 ============
ws3 = wb.create_sheet('變更1與變更2關係 (34筆)')
# 說明行
ws3.merge_cells('A1:H1')
ws3.cell(row=1, column=1,
    value='BSOB 34筆在 O-0002 中全部標為 X（不涉及），原因為 BSOB。變更1與變更2為獨立變更案。').font = Font(bold=True, color='CC0000')
ws3.cell(row=1, column=1).fill = note_fill

headers = [
    '料號', '品名', '封裝', 'Function',
    'O-0001 涉及與否', 'O-0001 AU類型', 'BSOB',
    'O-0002 涉及與否', 'O-0002 不涉及原因(推斷)'
]
for j, h in enumerate(headers, 1):
    ws3.cell(row=2, column=j, value=h)
style_header(ws3, 2, hdr_fill2, len(headers))

row_num = 3
for p in sorted(change1_parts):
    c1 = change1_dict.get(p, (None,)*10)
    c2_info = change2_all_dict.get(p, {})
    vals = [
        p,
        c1[1] if len(c1) > 1 and c1[1] else '',
        c1[2] if len(c1) > 2 and c1[2] else '',
        c1[3] if len(c1) > 3 and c1[3] else '',
        c1[4] if len(c1) > 4 and c1[4] else '',
        c1[5] if len(c1) > 5 and c1[5] else '',
        c1[6] if len(c1) > 6 and c1[6] else '',
        c2_info.get('involvement', '不在O-0002中'),
        'BSOB（已轉鈀金銅線，不需再轉銅線）' if c2_info.get('involvement') == 'X' else '',
    ]
    for j, val in enumerate(vals, 1):
        c = ws3.cell(row=row_num, column=j, value=val)
        c.border = thin_border
    row_num += 1

auto_width(ws3, len(headers), row_num)
ws3.freeze_panes = 'A3'

# ============ Sheet 4: 僅變更3 ============
ws4 = wb.create_sheet(f'僅變更3 ({len(c3_only)}筆)')
headers = [
    '料號', '封裝', 'Wafer Type', 'Die Size', 'Die Attach',
    '處置(Sn/Ag)', 'Epoxy', 'Eutectic',
    'NPI Command', 'Au/Con', 'Lead Frame', 'PKG Code',
    '2024出貨(K)', '2025出貨(K)', '出貨加總'
]
for j, h in enumerate(headers, 1):
    ws4.cell(row=1, column=j, value=h)
style_header(ws4, 1, hdr_fill3, len(headers))

row_num = 2
for p in sorted(c3_only):
    c3 = change3_dict.get(p, {})
    vals = [
        p, c3.get('package', ''), c3.get('wafer_type', ''), c3.get('die_size', ''),
        c3.get('die_attach', ''), c3.get('sn_ag', ''), c3.get('epoxy', ''), c3.get('eutectic', ''),
        c3.get('npi_command', ''), c3.get('au_con', ''), c3.get('lead_frame', ''), c3.get('pkg_code', ''),
        c3.get('ship_2024', 0), c3.get('ship_2025', 0), c3.get('ship_total', 0),
    ]
    for j, val in enumerate(vals, 1):
        c = ws4.cell(row=row_num, column=j, value=val)
        c.border = thin_border
    row_num += 1

auto_width(ws4, len(headers), row_num)
ws4.freeze_panes = 'A2'

# ============ Sheet 5: 僅變更2(V) ============
ws5 = wb.create_sheet(f'僅變更2(V) ({len(c2_only)}筆)')
headers = [
    '料號', '品名', '封裝', 'Function',
    '涉及與否', 'AU類型', '評估批次', '有BOM',
    'BOP金', 'BOP銀', 'BOP銅', '線材'
]
for j, h in enumerate(headers, 1):
    ws5.cell(row=1, column=j, value=h)
style_header(ws5, 1, hdr_fill5, len(headers))

row_num = 2
for p in sorted(c2_only):
    c2 = change2_dict.get(p, (None,)*12)
    vals = [
        p,
        c2[1] if len(c2) > 1 and c2[1] else '',
        c2[2] if len(c2) > 2 and c2[2] else '',
        c2[3] if len(c2) > 3 and c2[3] else '',
        c2[4] if len(c2) > 4 and c2[4] else '',
        c2[5] if len(c2) > 5 and c2[5] else '',
        c2[6] if len(c2) > 6 and c2[6] else '',
        c2[7] if len(c2) > 7 and c2[7] else '',
        c2[8] if len(c2) > 8 and c2[8] else '',
        c2[9] if len(c2) > 9 and c2[9] else '',
        c2[10] if len(c2) > 10 and c2[10] else '',
        c2[11] if len(c2) > 11 and c2[11] else '',
    ]
    for j, val in enumerate(vals, 1):
        c = ws5.cell(row=row_num, column=j, value=val)
        c.border = thin_border
    row_num += 1

auto_width(ws5, len(headers), row_num)
ws5.freeze_panes = 'A2'

# ============ Sheet 6: 全料號主表 ============
ws6 = wb.create_sheet('全料號參與對照')
headers = ['料號', '變更1(BSOB)', '變更2(Cu線,V)', '變更3(背錫銀)', '參與變更數', '跨變更']
for j, h in enumerate(headers, 1):
    ws6.cell(row=1, column=j, value=h)
style_header(ws6, 1, hdr_fill4, len(headers))

all_parts = sorted(all_union)
row_num = 2
for p in all_parts:
    in1 = 'V' if p in change1_parts else ''
    in2 = 'V' if p in change2_parts else ''
    in3 = 'V' if p in change3_parts else ''
    count = (1 if p in change1_parts else 0) + (1 if p in change2_parts else 0) + (1 if p in change3_parts else 0)
    cross = 'Y' if count > 1 else ''
    vals = [p, in1, in2, in3, count, cross]
    for j, val in enumerate(vals, 1):
        c = ws6.cell(row=row_num, column=j, value=val)
        c.border = thin_border
        if j > 1:
            c.alignment = Alignment(horizontal='center')
        if count > 1:
            c.fill = overlap_fill
    row_num += 1

auto_width(ws6, len(headers), row_num)
ws6.freeze_panes = 'A2'
ws6.auto_filter.ref = f'A1:F{row_num - 1}'

# === Save ===
output_path = str(_ROOT / 'projects' / 'ecr-ecn' / 'vault' / 'outputs' / '跨變更交叉分析報告.xlsx')
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f'\nDone! Saved to: {output_path}')
print(f'Sheets: {wb.sheetnames}')
print(f'Sheet row counts:')
for name in wb.sheetnames:
    ws = wb[name]
    print(f'  {name}: {ws.max_row} rows')
