"""共用 Excel 樣式定義 — 供 Python 腳本（openpyxl）和 MCP 工具共用。

所有報告腳本的色碼、字型、預設樣式組合統一在此定義。
修改色碼只需改此檔，所有報告自動同步。
"""

# ── 企業標準色（HEX） ──
COLORS = {
    "header_bg": "1F4E79",
    "header_fg": "FFFFFF",
    "subheader_bg": "D6E4F0",
    "confirm": "C6EFCE",       # 綠
    "warning": "FFEB9C",       # 黃
    "error": "FFC7CE",         # 紅
    "accent": "F4B084",        # 橘
    "gray_alt": "F2F2F2",      # 交替行灰
    "light_blue": "DAEEF3",    # 淺藍
    "title_color": "1F4E79",   # 標題文字色
    "note_color": "666666",    # 註釋文字色
    "alert_color": "C00000",   # 警示文字色
}

# ── 字體預設 ──
FONTS = {
    "header":    {"name": "Calibri", "size": 10, "bold": True,  "color": "FFFFFF"},
    "subheader": {"name": "Calibri", "size": 10, "bold": True},
    "data":      {"name": "Calibri", "size": 9,  "bold": False},
    "data_bold": {"name": "Calibri", "size": 9,  "bold": True},
    "title":     {"name": "Calibri", "size": 14, "bold": True,  "color": "1F4E79"},
    "subtitle":  {"name": "Calibri", "size": 10, "bold": True,  "color": "2E75B6"},
    "section":   {"name": "Calibri", "size": 11, "bold": True,  "color": "1F4E79"},
    "note":      {"name": "Calibri", "size": 9,  "italic": True, "color": "666666"},
}

# ── 樣式組合（供 MCP apply_style_preset 使用）──
PRESETS = {
    "header": {
        "bg_color": "1F4E79", "font_color": "FFFFFF",
        "font_name": "Calibri", "font_bold": True, "font_size": 10,
        "h_align": "center", "v_align": "center", "wrap_text": True,
        "border": "thin",
    },
    "subheader": {
        "bg_color": "D6E4F0",
        "font_name": "Calibri", "font_bold": True, "font_size": 10,
        "h_align": "center", "v_align": "center", "wrap_text": True,
        "border": "thin",
    },
    "data": {
        "font_name": "Calibri", "font_size": 9,
        "v_align": "center", "wrap_text": True,
        "border": "thin",
    },
    "data_alt": {
        "bg_color": "F2F2F2",
        "font_name": "Calibri", "font_size": 9,
        "v_align": "center", "wrap_text": True,
        "border": "thin",
    },
    "confirm": {"bg_color": "C6EFCE", "font_size": 9, "border": "thin"},
    "warning": {"bg_color": "FFEB9C", "font_size": 9, "border": "thin"},
    "error":   {"bg_color": "FFC7CE", "font_size": 9, "border": "thin"},
    "accent":  {"bg_color": "F4B084", "font_size": 9, "border": "thin"},
}


# ── openpyxl 適配器 ──
def to_openpyxl_styles():
    """將共用定義轉為 openpyxl 樣式物件字典。

    Returns:
        dict with keys: fills, fonts, border, alignment
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    fills = {
        "header":    PatternFill("solid", fgColor=COLORS["header_bg"]),
        "subheader": PatternFill("solid", fgColor=COLORS["subheader_bg"]),
        "confirm":   PatternFill("solid", fgColor=COLORS["confirm"]),
        "warning":   PatternFill("solid", fgColor=COLORS["warning"]),
        "error":     PatternFill("solid", fgColor=COLORS["error"]),
        "accent":    PatternFill("solid", fgColor=COLORS["accent"]),
        "gray_alt":  PatternFill("solid", fgColor=COLORS["gray_alt"]),
        "light_blue": PatternFill("solid", fgColor=COLORS["light_blue"]),
    }

    fonts = {}
    for key, spec in FONTS.items():
        kwargs = {"name": spec.get("name", "Calibri")}
        if "size" in spec:
            kwargs["size"] = spec["size"]
        if spec.get("bold"):
            kwargs["bold"] = True
        if spec.get("italic"):
            kwargs["italic"] = True
        if "color" in spec:
            kwargs["color"] = spec["color"]
        fonts[key] = Font(**kwargs)

    thin_side = Side(style="thin")
    border = Border(left=thin_side, right=thin_side,
                    top=thin_side, bottom=thin_side)

    alignments = {
        "header": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "data":   Alignment(vertical="center", wrap_text=True),
        "center": Alignment(horizontal="center"),
    }

    return {
        "fills": fills,
        "fonts": fonts,
        "border": border,
        "alignments": alignments,
    }


def style_header_row(ws, row, max_col):
    """套用標題列樣式（openpyxl worksheet）。"""
    s = to_openpyxl_styles()
    for ci in range(1, max_col + 1):
        cell = ws.cell(row, ci)
        cell.fill = s["fills"]["header"]
        cell.font = s["fonts"]["header"]
        cell.alignment = s["alignments"]["header"]
        cell.border = s["border"]


def style_data_row(ws, row, max_col, fill=None, bold_cols=None):
    """套用資料列樣式（openpyxl worksheet）。

    Args:
        fill: openpyxl PatternFill（可選，用於條件行色）
        bold_cols: list of 1-based column indices to bold
    """
    s = to_openpyxl_styles()
    for ci in range(1, max_col + 1):
        cell = ws.cell(row, ci)
        if bold_cols and ci in bold_cols:
            cell.font = s["fonts"]["data_bold"]
        else:
            cell.font = s["fonts"]["data"]
        if fill:
            cell.fill = fill
        cell.alignment = s["alignments"]["data"]
        cell.border = s["border"]
