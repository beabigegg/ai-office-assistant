#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
MES Report Downloader — SSRS URL Access

Two authentication modes:
  - ntlm (default): requests + NTLM, requires VPN  [recommended]
  - playwright: Chromium + NTLM, also requires VPN (external DNS removed by IT, 2026-03-30)

Usage:
    python shared/tools/mes_report.py list
    python shared/tools/mes_report.py download PJMES002 --start 2026/03/01 --end 2026/03/19
    python shared/tools/mes_report.py download PJMES002 --start 2026/03/01 --end 2026/03/19 --auth playwright
    python shared/tools/mes_report.py download PJMES002 --start 2026/03/01 --end 2026/03/19 --auth playwright --merge
    python shared/tools/mes_report.py params PJMES002
    python shared/tools/mes_report.py test [--auth playwright]
"""

import argparse
import os
import re
import sys
import time
from calendar import monthrange
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

import requests
import urllib3
import yaml
from requests_ntlm import HttpNtlmAuth

# Load .env
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).resolve().parent.parent.parent / ".env"
    if _env_path.exists():
        load_dotenv(_env_path)
except ImportError:
    pass

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── Paths ─────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent.parent
CONFIG_PATH = ROOT / "projects" / "mes-report" / "workspace" / "mes_reports.yaml"
OUTPUT_DIR = ROOT / "projects" / "mes-report" / "vault" / "outputs"


# ── Credentials ──────────────────────────────────────
def _get_credentials():
    user = os.environ.get("MES_USER", "")
    pwd = os.environ.get("MES_PASSWORD", "")
    if not user or not pwd:
        print("ERROR: Set MES_USER and MES_PASSWORD in .env")
        sys.exit(1)
    return user, pwd


def get_auth():
    user, pwd = _get_credentials()
    return HttpNtlmAuth(user, pwd)


# ── Config ───────────────────────────────────────────
def load_config(config_path=None):
    path = Path(config_path) if config_path else CONFIG_PATH
    if not path.exists():
        print(f"ERROR: Config not found: {path}")
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


# ── Date splitting ───────────────────────────────────
def split_date_ranges(start_str, end_str, chunk_days, chunk_mode, date_fmt):
    start = datetime.strptime(start_str, "%Y/%m/%d")
    end = datetime.strptime(end_str, "%Y/%m/%d")
    if start >= end:
        print(f"ERROR: start {start_str} >= end {end_str}")
        return []

    ranges = []
    if chunk_mode == "month":
        cur = start
        last_overall = end - timedelta(days=1)
        while cur <= last_overall:
            days_in_month = monthrange(cur.year, cur.month)[1]
            last_day = datetime(cur.year, cur.month, days_in_month)
            seg_end = min(last_day, last_overall)
            ranges.append((_fmt_date(cur, date_fmt), _fmt_date(seg_end, date_fmt)))
            cur = last_day + timedelta(days=1)
    else:
        cur = start
        while cur < end:
            nxt = min(cur + timedelta(days=chunk_days), end)
            ranges.append((_fmt_date(cur, date_fmt), _fmt_date(nxt, date_fmt)))
            cur = nxt
    return ranges


def _fmt_date(dt, fmt):
    if fmt == "YYYYMMDD HH24MISS":
        return dt.strftime("%Y%m%d 000000")
    return dt.strftime("%Y%m%d")


# ── URL builder ──────────────────────────────────────
def _build_url(base_url, report_path, params):
    encoded_path = quote(report_path, safe="/ ")
    parts = [f"{base_url}?{encoded_path}", "rs:Format=EXCELOPENXML"]
    for k, v in params.items():
        if v is None:
            continue
        if isinstance(v, list):
            for val in v:
                parts.append(f"{k}={quote(str(val), safe='')}")
        else:
            parts.append(f"{k}={quote(str(v), safe='')}")
    return "&".join(parts)


# ── SOAP valid values ────────────────────────────────
def _parse_valid_values_xml(xml_text):
    result = {}
    for block in re.split(r"<ReportParameter>", xml_text)[1:]:
        name_m = re.search(r"<Name>([^<]+)</Name>", block)
        if not name_m:
            continue
        vs = re.search(r"<ValidValues>(.*?)</ValidValues>", block, re.DOTALL)
        if vs:
            vals = re.findall(r"<Value>([^<]*)</Value>", vs.group(1))
            if vals:
                result[name_m.group(1)] = vals
    return result


_SOAP_BODY = """<?xml version="1.0" encoding="utf-8"?>
<soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/"
               xmlns:re="http://schemas.microsoft.com/sqlserver/2005/06/30/reporting/reportingservices">
  <soap:Body><re:LoadReport><re:Report>{path}</re:Report></re:LoadReport></soap:Body>
</soap:Envelope>"""

_SOAP_ACTION = "http://schemas.microsoft.com/sqlserver/2005/06/30/reporting/reportingservices/LoadReport"


def _fetch_valid_values_ntlm(session, base_url, report_path):
    soap_url = base_url.replace("/ReportServer", "/ReportServer/ReportExecution2005.asmx")
    try:
        r = session.post(soap_url,
                         data=_SOAP_BODY.format(path=report_path).encode("utf-8"),
                         headers={"Content-Type": "text/xml; charset=utf-8", "SOAPAction": _SOAP_ACTION},
                         timeout=15)
        return _parse_valid_values_xml(r.text) if r.status_code == 200 else {}
    except Exception:
        return {}


def _fetch_valid_values_pw(page, base_url, report_path):
    soap_url = base_url.replace("/ReportServer", "/ReportServer/ReportExecution2005.asmx")
    body = _SOAP_BODY.format(path=report_path)
    try:
        xml = page.evaluate("""async ([url, body, action]) => {
            const r = await fetch(url, {
                method: 'POST',
                headers: {'Content-Type':'text/xml; charset=utf-8','SOAPAction':action},
                body: body
            });
            return await r.text();
        }""", [soap_url, body, _SOAP_ACTION])
        return _parse_valid_values_xml(xml)
    except Exception:
        return {}


# ── NTLM download (requests) ────────────────────────
def _dl_ntlm(session, base_url, report_path, params, timeout=300):
    url = _build_url(base_url, report_path, params)
    try:
        r = session.get(url, timeout=(30, timeout), stream=True)
    except requests.exceptions.ReadTimeout:
        print("  TIMEOUT")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"  Connection error: {e}")
        return None

    ct = r.headers.get("Content-Type", "")
    ok = any(x in ct.lower() for x in ["spreadsheet", "openxml", "octet-stream"])
    if r.status_code != 200 or not ok:
        print(f"  HTTP {r.status_code}, CT: {ct}")
        if r.status_code != 200:
            return None
        text = r.text[:500] if hasattr(r, "text") else ""
        if "rsReportParameter" in text:
            print("  SSRS param error")
            return None

    data = b""
    for chunk in r.iter_content(chunk_size=8192):
        data += chunk
    return data


# ── Playwright download ─────────────────────────────
def _dl_pw(page, base_url, report_path, params, save_path, timeout=300):
    url = _build_url(base_url, report_path, params)
    try:
        with page.expect_download(timeout=timeout * 1000) as dl_info:
            page.evaluate("(url) => { window.location.href = url; }", url)
        dl_info.value.save_as(str(save_path))
        return os.path.getsize(save_path)
    except Exception as e:
        print(f"  Download error: {e}")
        return None


# ── Merge Excel ──────────────────────────────────────
def merge_excels(file_list, output_path):
    import openpyxl
    if len(file_list) == 1:
        os.rename(file_list[0], output_path)
        return

    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
    header_written = False

    for fpath in file_list:
        wb = openpyxl.load_workbook(fpath, read_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                if not header_written:
                    merged_ws.append(row)
                continue
            if i == 1:
                if not header_written:
                    merged_ws.append(row)
                    header_written = True
                continue
            merged_ws.append(row)
        wb.close()

    merged_wb.save(output_path)
    merged_wb.close()
    for fpath in file_list:
        os.remove(fpath)


# ══════════════════════════════════════════════════════
# CLI Commands
# ══════════════════════════════════════════════════════

def cmd_list(args):
    config = load_config(args.config)
    print(f"{'ID':<12} {'Name':<40} {'Chunk':<8} {'Mode':<6}")
    print("-" * 70)
    for rid, rdef in config["reports"].items():
        enabled = rdef.get("enabled", True)
        chunk = rdef.get("date_chunk_days", "?")
        mode = rdef.get("date_chunk_mode", "day")
        chunk_str = f"{chunk}d" if mode == "day" else "month"
        status = "" if enabled else " [disabled]"
        print(f"{rid:<12} {rdef['name']:<40} {chunk_str:<8} {mode:<6}{status}")


def cmd_params(args):
    config = load_config(args.config)
    rid = args.report_id.upper()
    if rid not in config["reports"]:
        print(f"ERROR: Unknown report {rid}")
        return
    rdef = config["reports"][rid]
    print(f"Report: {rid} - {rdef['name']}")
    print(f"Path: {rdef['path']}")
    print(f"Chunk: {rdef.get('date_chunk_days', '?')}d ({rdef.get('date_chunk_mode', 'day')})")
    print(f"\n{'Param':<30} {'Role/Default':<30} {'Null'}")
    print("-" * 65)
    for pname, pdef in rdef["params"].items():
        val = f"[{pdef['role']}]" if "role" in pdef else repr(pdef.get("default", ""))
        null = "Y" if pdef.get("nullable") else ""
        print(f"{pname:<30} {val:<30} {null}")


def cmd_download(args):
    config = load_config(args.config)
    rid = args.report_id.upper()
    if rid not in config["reports"]:
        print(f"ERROR: Unknown report {rid}")
        return

    rdef = config["reports"][rid]
    if not rdef.get("enabled", True):
        print(f"WARNING: {rid} disabled")
        return

    base_url = config["ssrs_base"]
    date_ranges = split_date_ranges(
        args.start, args.end,
        rdef.get("date_chunk_days", 30),
        rdef.get("date_chunk_mode", "day"),
        rdef.get("date_format", "YYYYMMDD"),
    )
    if not date_ranges:
        return

    print(f"Report: {rid} - {rdef['name']}")
    print(f"Range: {args.start} ~ {args.end} -> {len(date_ranges)} segments")
    print(f"Auth: {args.auth}")

    # Build params
    base_params = {}
    select_all_params = []
    for pname, pdef in rdef["params"].items():
        if "role" not in pdef:
            if pdef.get("select_all"):
                select_all_params.append(pname)
            else:
                val = pdef.get("default", "")
                if pdef.get("nullable") and val == "":
                    base_params[pname + ":isnull"] = "true"
                else:
                    base_params[pname] = val

    if args.param:
        for p in args.param:
            k, v = p.split("=", 1)
            base_params[k] = v

    start_param = end_param = None
    for pname, pdef in rdef["params"].items():
        if pdef.get("role") == "start_date":
            start_param = pname
        elif pdef.get("role") == "end_date":
            end_param = pname

    out_dir = Path(args.output) if args.output else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)

    # Download
    if args.auth == "playwright":
        files = _loop_pw(config, rdef, rid, base_url, date_ranges,
                         base_params, select_all_params, start_param, end_param, out_dir, args)
    else:
        files = _loop_ntlm(config, rdef, rid, base_url, date_ranges,
                           base_params, select_all_params, start_param, end_param, out_dir, args)

    if not files:
        print("No files downloaded")
        return

    if args.merge and len(files) > 1:
        merged = out_dir / f"{rid}_{args.start.replace('/', '')}_{args.end.replace('/', '')}_merged.xlsx"
        print(f"Merging {len(files)} files ... ", end="", flush=True)
        merge_excels(files, str(merged))
        print(f"OK -> {merged}")
    else:
        print(f"Done: {len(files)} files in {out_dir}")


def _loop_ntlm(config, rdef, rid, base_url, date_ranges,
               base_params, select_all_params, start_param, end_param, out_dir, args):
    auth = get_auth()
    session = requests.Session()
    session.auth = auth
    session.verify = False
    try:
        session.get(base_url, timeout=15)
    except Exception:
        pass

    if select_all_params:
        vv = _fetch_valid_values_ntlm(session, base_url, rdef["path"])
        for pname in select_all_params:
            if pname in vv:
                base_params[pname] = vv[pname]
                print(f"  {pname}: {len(vv[pname])} values")
            else:
                base_params[pname] = ""

    files = []
    for i, (sd, ed) in enumerate(date_ranges, 1):
        params = dict(base_params)
        params[start_param] = sd
        params[end_param] = ed
        label = f"{sd.split()[0]}~{ed.split()[0]}"
        print(f"  [{i}/{len(date_ranges)}] {label} ... ", end="", flush=True)

        timeout = rdef.get("timeout", config.get("default_timeout", args.timeout))
        data = _dl_ntlm(session, base_url, rdef["path"], params, timeout=timeout)
        if data is None:
            print("FAILED")
            continue

        fname = f"{rid}_{sd.replace(' ', '_')}_{ed.replace(' ', '_')}.xlsx"
        fpath = out_dir / fname
        with open(fpath, "wb") as f:
            f.write(data)
        print(f"OK ({len(data):,} bytes)")
        files.append(str(fpath))
        if i < len(date_ranges):
            time.sleep(2)
    return files


def _loop_pw(config, rdef, rid, base_url, date_ranges,
             base_params, select_all_params, start_param, end_param, out_dir, args):
    from playwright.sync_api import sync_playwright

    user, pwd = _get_credentials()
    files = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            ignore_https_errors=True,
            http_credentials={"username": user, "password": pwd},
        )
        page = context.new_page()

        # Warm up: establish NTLM session through Access Gateway
        print("  Connecting ... ", end="", flush=True)
        try:
            resp = page.goto(base_url, wait_until="domcontentloaded", timeout=30000)
            if resp and resp.status == 200:
                print("OK")
            else:
                print(f"HTTP {resp.status if resp else '?'}")
                browser.close()
                return files
        except Exception as e:
            print(f"Error: {e}")
            browser.close()
            return files

        # select_all params
        if select_all_params:
            vv = _fetch_valid_values_pw(page, base_url, rdef["path"])
            for pname in select_all_params:
                if pname in vv:
                    base_params[pname] = vv[pname]
                    print(f"  {pname}: {len(vv[pname])} values")
                else:
                    base_params[pname] = ""

        # Download loop
        for i, (sd, ed) in enumerate(date_ranges, 1):
            params = dict(base_params)
            params[start_param] = sd
            params[end_param] = ed
            label = f"{sd.split()[0]}~{ed.split()[0]}"
            print(f"  [{i}/{len(date_ranges)}] {label} ... ", end="", flush=True)

            fname = f"{rid}_{sd.replace(' ', '_')}_{ed.replace(' ', '_')}.xlsx"
            fpath = out_dir / fname
            timeout = rdef.get("timeout", config.get("default_timeout", args.timeout))
            size = _dl_pw(page, base_url, rdef["path"], params, str(fpath), timeout=timeout)

            if size is None:
                print("FAILED")
                continue

            print(f"OK ({size:,} bytes)")
            files.append(str(fpath))
            if i < len(date_ranges):
                time.sleep(2)

        browser.close()
    return files


def cmd_test(args):
    config = load_config(args.config)
    base_url = config["ssrs_base"]

    if args.auth == "playwright":
        from playwright.sync_api import sync_playwright
        user, pwd = _get_credentials()
        print(f"Test (playwright): {base_url}")
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                ignore_https_errors=True,
                http_credentials={"username": user, "password": pwd},
            )
            page = ctx.new_page()
            try:
                resp = page.goto(base_url, wait_until="domcontentloaded", timeout=30000)
                if resp and resp.status == 200:
                    print(f"OK (HTTP {resp.status})")
                else:
                    print(f"FAIL (HTTP {resp.status if resp else '?'})")
            except Exception as e:
                print(f"Error: {e}")
            browser.close()
    else:
        auth = get_auth()
        print(f"Test (ntlm): {base_url}")
        try:
            r = requests.get(base_url, auth=auth, verify=False, timeout=15)
            if r.status_code == 200:
                print(f"OK (HTTP {r.status_code})")
            else:
                print(f"FAIL (HTTP {r.status_code})")
        except Exception as e:
            print(f"Error: {e}")


# ── Main ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="MES Report Downloader")
    parser.add_argument("--config", help="Config file path")
    parser.add_argument("--auth", choices=["ntlm", "playwright"], default="ntlm",
                        help="ntlm (VPN) or playwright (no VPN)")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("list", help="List reports")

    p_p = sub.add_parser("params", help="Show report params")
    p_p.add_argument("report_id")

    p_dl = sub.add_parser("download", help="Download report")
    p_dl.add_argument("report_id")
    p_dl.add_argument("--start", required=True, help="Start date (YYYY/MM/DD)")
    p_dl.add_argument("--end", required=True, help="End date (YYYY/MM/DD)")
    p_dl.add_argument("--merge", action="store_true", help="Merge segments")
    p_dl.add_argument("--param", action="append", help="Override param (KEY=VALUE)")
    p_dl.add_argument("--output", help="Output directory")
    p_dl.add_argument("--timeout", type=int, default=300, help="Segment timeout (s)")

    sub.add_parser("test", help="Test connection")

    args = parser.parse_args()
    if args.command == "list":
        cmd_list(args)
    elif args.command == "params":
        cmd_params(args)
    elif args.command == "download":
        cmd_download(args)
    elif args.command == "test":
        cmd_test(args)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
